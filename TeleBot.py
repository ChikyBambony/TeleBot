import os
import datetime
import telebot
import loging
from config import get_settings

st = get_settings()
bot = telebot.TeleBot(st.telegram_bot_token)
users = {}

def _is_valid_phone(text: str) -> bool:
    return (
        len(text) == 10
        and text.isdigit()
        and text[0] == "9"
        and text[1] != "4"
    )

@bot.message_handler(content_types=["text"])
def on_text(message):
    text = (message.text or "").strip()

    if message.chat.id in st.blacklist_ids:
        bot.send_message(message.chat.id, "Подозрительные действия пользователя, функции чата отключены")
        return

    if text == "/start":
        bot.send_message(message.chat.id, "Привет! Отправь номер (10 цифр), пример: 9048632908")
        return

    # Админы: запрос в BDPN
    if message.chat.id in st.admin_ids and "/" not in text:
        if not _is_valid_phone(text):
            bot.send_message(message.chat.id, "Номер некорректен (нужно 10 цифр, начинается с 9)")
            return
        try:
            answer = loging.login(text)
        except Exception as e:
            answer = f"Ошибка при запросе к серверу: {e}"
        bot.send_message(message.chat.id, answer)
        return

    # Не админы: Excel/логи только если настроено
    excel_path = os.getenv("EXCEL_PATH", "").strip()
    excel_sheet = os.getenv("EXCEL_SHEET", "").strip()
    log_actions = os.getenv("LOG_ACTIONS_PATH", "").strip()
    log_ids = os.getenv("LOG_IDS_PATH", "").strip()

    if not excel_path or not log_actions or not log_ids:
        now = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
        bot.send_message(message.chat.id, f"Доступ ограничен. (не настроены пути EXCEL/LOG в .env) ({now})")
        return

    # Импорты только если реально используем Excel
    import openpyxl as xl
    import pandas as pd

    users.update({message.from_user.username: message.chat.id})
    now_dt = datetime.datetime.now()
    buff_request = text

    # Логи — аккуратно через with
    with open(log_actions, "a", encoding="utf-8") as f, open(log_ids, "a", encoding="utf-8") as d:
        f.write(f"{message.from_user.first_name}  {buff_request}")
        d.write(f"{message.from_user.first_name} - {message.chat.id}\n")

        wb = xl.load_workbook(excel_path)
        ws = wb[excel_sheet]
        df = pd.read_excel(excel_path, 0)

        i = 2
        k = len(df) + 1

        for o in range(i, k):
            if ws.cell(row=o, column=2).value == buff_request:
                if ws.cell(row=o, column=16).value:
                    buf = f"{ws.cell(row=o, column=4).value}, Ответ от нас: {ws.cell(row=o, column=16).value}"
                    bot.send_message(message.chat.id, buf)
                    f.write(" - " + buf + now_dt.strftime(" %d-%m-%Y %H:%M") + "\n")
                    return
                else:
                    buf = str(ws.cell(row=o, column=4).value)
                    bot.send_message(message.chat.id, buf)
                    f.write(" - " + buf + now_dt.strftime(" %d-%m-%Y %H:%M") + "\n")
                    return

        bot.send_message(message.chat.id, "Данных нет")
        f.write(" - Данных нет" + now_dt.strftime(" %d-%m-%Y %H:%M\n"))

bot.infinity_polling(skip_pending=True)
